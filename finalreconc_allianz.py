import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from itertools import combinations
import warnings
import warnings
warnings.filterwarnings("ignore")
warnings.filterwarnings("default")
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase  
from email import encoders
import os




#print("def1")
def1=None
def2=None


def Process_Files(def1,def2,def3):
    df_bank=pd.read_excel(def1)

    pd.set_option('display.float_format',lambda x : '{:,.2f}'.format(x))
    # Use the existing file name as-is
    base_filename = os.path.basename(def1)

    # Extract the file name without extension
    file_name_without_extension, _ = os.path.splitext(base_filename)


    file_name_without_extension
    full_name=f"RECONCILED/{file_name_without_extension}_Recon.xlsx"
    print(full_name)
    df_bank.columns = df_bank.columns.str.strip()
    df_bank.dropna(subset="Running Balance",inplace=True)
    df_bank=df_bank.rename(columns={'Transaction Date':'Date'})
    df_bank=df_bank.rename(columns={'Transaction Type':'VoucherN°'})
    df_bank=df_bank.rename(columns={'Transaction Details':'Description'})
    df_bank=df_bank.rename(columns={'Debits':'Direct Debits'})
    df_bank=df_bank.rename(columns={'Credits':'Un receipted Items'})
    df_bank.dropna(subset="Running Balance",inplace=True)
    df_bank=df_bank[['Date', 'VoucherN°', 'Description','Direct Debits', 'Un receipted Items']]
    cashbook=pd.read_excel(def2)
    cashbook.dropna(subset="Journal Name",inplace=True)
    cashbook=cashbook.rename(columns={'Transaction Date':'Date'})
    cashbook=cashbook.rename(columns={'Reference Date':'Cheque N°'})
    cashbook=cashbook.rename(columns={'Line Description':'Description'})

    if "USD" in file_name_without_extension:
        Opening_balance= 220153.35
        float_columns = ['Debit FC', 'Credit FC']
        cashbook[float_columns] = cashbook[float_columns].astype(float)
        Total_Debit_FC = cashbook['Debit FC'].sum().round(2)
        Total_Credit_FC = cashbook['Credit FC'].sum().round(2)
        closing_balance_FC=(Opening_balance+Total_Debit_FC-Total_Credit_FC).round(2)
        print(closing_balance_FC)
        print(Total_Debit_FC)
        print(Total_Credit_FC)
        cashbook=cashbook.rename(columns={'Debit FC':'Un credited Items'})
        cashbook=cashbook.rename(columns={'Credit FC':'Un Paid Items'})
    else:
        cashbook=cashbook.rename(columns={'Debit LC':'Un credited Items'})
        cashbook=cashbook.rename(columns={'Credit LC':'Un Paid Items'})    
    cashbook=cashbook[['Date', 'Cheque N°', 'Description','Un credited Items', 'Un Paid Items']]
    cashbook.shape
    previous_workings=pd.read_excel(def3)
    previous_workings.columns = previous_workings.columns.str.strip()
    previous_workings.shape
    num_columns=['Direct Debits', 'Un receipted Items', 'Un credited Items','Un Paid Items']


    num_columns=['Direct Debits', 'Un receipted Items', 'Un credited Items','Un Paid Items']

    workings1=pd.concat([df_bank,cashbook],ignore_index=True)
    



    for col in num_columns:
        workings1[col]=workings1[col].astype(str)
        workings1[col] = pd.to_numeric(workings1[col].replace('-', pd.NaT).str.replace(',', ''), errors='coerce').fillna(0)

    workings1['Matching']=workings1['Direct Debits']+workings1['Un Paid Items']+workings1['Un credited Items']+workings1['Un receipted Items']
    workings1
    workings=pd.concat([workings1,previous_workings])
    workings
    workings=workings[['Date', 'VoucherN°', 'Cheque N°', 'Description','Direct Debits', 'Un receipted Items', 'Un credited Items','Un Paid Items','Matching']]
    num_columns=['Direct Debits', 'Un receipted Items', 'Un credited Items','Un Paid Items','Matching']
    for col in num_columns:
        workings[col]=workings[col].astype(str)
        workings[col] = pd.to_numeric(workings[col].replace('-', pd.NaT).str.replace(',', ''), errors='coerce').fillna(0)
    workings=workings.sort_values(by='Matching',ascending=False)
    workings.reset_index(drop=True,inplace=True)
    workings['Indexing']=range(1,len(workings)+1)
    workings

    # Assuming workings is your DataFrame
    pd.set_option('display.float_format', '{:.2f}'.format)

    # Now display your DataFrame

    workings['Amount_A']=workings['Direct Debits']+workings['Un receipted Items']

    workings['Amount_B']=workings['Un credited Items']+workings['Un Paid Items']

    workings['Amount_B'] = workings['Amount_B'].apply(lambda x: round(x) if x > 1 else x)
    workings['Amount_A'] = workings['Amount_A'].apply(lambda x: round(x) if x > 1 else x)

    df_A=workings[workings['Amount_B']<1]
    df_B=workings[workings['Amount_A']<1]
    df_A.reset_index(drop=True,inplace=True)
    df_B.reset_index(drop=True,inplace=True)


    df_A=df_A[['Date','VoucherN°','Cheque N°','Description','Direct Debits','Un receipted Items','Un credited Items','Un Paid Items','Matching','Amount_A','Indexing']]
    df_B=df_B[['Date','VoucherN°','Cheque N°','Description','Direct Debits','Un receipted Items','Un credited Items','Un Paid Items','Matching','Amount_B','Indexing']]

    df_A.to_csv('dfa.csv')
    df_B.to_csv('dfb.csv')
    for index,row in df_A.iterrows():

        df_temp=df_B[(df_B['Amount_B']== row['Amount_A'] )& (df_B['Amount_B'] > 0)]

        if df_temp.shape[0] > 0:
            if row['Amount_A']==0.12:
                print('no')

        
            index_remove=df_B[df_B['Amount_B']==row['Amount_A']].index[0]
            df_B.drop(index_remove,inplace=True)
            df_A.drop(index,inplace=True)


        # else:
        #     if row['Amount_A']==250000:
        #         print('no')
        #     # print(df_temp.head(5))
        #     df_unreconcileA=pd.concat([df_unreconcileA,df_tempb],ignore_index=True)
        #     #df_reconcile.append(df_temp,ignore_index=True)

    unreconciled=pd.concat([df_A,df_B],ignore_index=True)
    unreconciled = unreconciled.drop_duplicates(subset='Indexing', keep='first')
    unreconciled.to_csv("allyn.csv")
    unreconciled.shape
    reconciled = workings[~workings['Indexing'].isin(unreconciled['Indexing'])]
    reconciled = reconciled.drop_duplicates(subset='Indexing', keep='first')

    reconciled.to_csv("recon.csv")
    reconciled.shape




    
    df_bank1=pd.read_excel(def1)
    df_bank1.columns = df_bank1.columns.str.strip()


    # Convert the 'Transaction Transaction Date' column to Transaction Datetime
    df_bank1['Transaction Date'] = pd.to_datetime(df_bank1['Transaction Date'], errors='coerce')

    # Sort the DataFrame by the ' Date' column
    df_bank1 = df_bank1.sort_values(by='Transaction Date',ascending=True)

    # If you want to reset the index after sorting
    df_bank1 = df_bank1.reset_index(drop=True)
    df_bank1.dropna(subset='Running Balance',inplace=True)

    # Print the sorted DataFrame

    running_balance_last_row = df_bank1['Running Balance'].iloc[-1]
    if isinstance(running_balance_last_row, str) and ',' in running_balance_last_row:
        running_balance_last_row = float(running_balance_last_row.replace(',', ''))
    elif isinstance(running_balance_last_row, (int, float)):
        running_balance_last_row = float(running_balance_last_row)
    cashbook1=pd.read_excel(def2)
    cashbook1.dropna(subset='Running Total',inplace=True)
    running_cashbook_last_row = cashbook1['Running Total'].iloc[-1]
    running_cashbook_last_row=running_cashbook_last_row
    # Assuming unreconciled is your non-reconciled dataframe
    '''unreconciled=unreconciled.dropna(subset='Description')

    sum_row['Date'] = 'Sub Total'

    #unreconciled.loc[len(unreconciled)] = sum_row
    unreconciled.reset_index(drop=True, inplace=True)'''
    # Convert sum_row to a DataFrame with a single row
    sum_row = unreconciled.select_dtypes(include='float').sum()

    sum_row_df = pd.DataFrame([sum_row], columns=sum_row.index)

    # Concatenate unreconciled and sum_row_df
    unreconciled = pd.concat([unreconciled, sum_row_df], ignore_index=True)

    # Set 'Date' to 'Sub Total' in the last row
    unreconciled.loc[len(unreconciled), 'Date'] = 'Sub Total'
    unreconciled=unreconciled.dropna(subset='Matching')

    unreconciled.reset_index(drop=True, inplace=True)
    cash_reconciled = cashbook[cashbook['Description'].isin(reconciled['Description'])]
    cash_nonreconciled = cashbook[~cashbook['Description'].isin(reconciled['Description'])]

    selected_columns = ['Date', 'VoucherN°', 'Cheque N°', 'Description', 'Direct Debits', 'Un receipted Items', 'Un credited Items', 'Un Paid Items','Matching']
    reconciled = reconciled[selected_columns]
    unreconciled = unreconciled[selected_columns]
    workings = workings[selected_columns]
    workings = workings[workings['Matching'] > 0]
    unreconciled = unreconciled[unreconciled['Matching'] > 0]

    sum1=unreconciled['Un credited Items'].iloc[-1]
    sum2=unreconciled['Direct Debits'].iloc[-1]
    sum3=unreconciled['Un receipted Items'].iloc[-1]
    sum4=unreconciled['Un Paid Items'].iloc[-1]

    # Load the workbook
    workbook = openpyxl.load_workbook("8E3E8A00.xlsx")

    # Select the desired worksheet (replace "Sheet1" with the actual sheet name)
    worksheet = workbook["Sheet1"]

    # Write sum1 to cell C8
    
    if "USD" in full_name:
        worksheet["C10"] = closing_balance_FC

    else:
        worksheet["C10"] = running_cashbook_last_row

    #worksheet["C10"] = running_cashbook_last_row
    worksheet["D8"] = running_balance_last_row
    worksheet["D13"] = sum1
    worksheet["C14"] = sum2
    worksheet["C17"] = sum3
    worksheet["D18"] = sum4
    #worksheet["C24"]=worksheet["C10"]-worksheet["C14"]+worksheet["C17"]
    #worksheet["C24"]=worksheet["C10"]-worksheet["C14"]+worksheet["C17"]

    # Save the changes
    workbook.save(full_name)


    # Load an existing workbook
    existing_workbook_path = full_name
    existing_workbook = load_workbook(existing_workbook_path)


    # Get or create the 'Working' sheet in the existing workbook
    working_sheet = existing_workbook.get_sheet_by_name('Working') if 'Working' in existing_workbook.sheetnames else existing_workbook.create_sheet(title='Working')
    working_sheet.sheet_properties.tabColor = 'FFA500'  # Hex color code for orange

    # Starting row number for the data in the 'Working' sheet
    start_row_working = 7

    # Write column headers to the first row of the 'Working' sheet
    for col_num, header in enumerate(workings.columns, 1):
        working_sheet.cell(row=start_row_working - 1, column=col_num, value=header)

    # Write the 'Working' DataFrame to the specified range in the 'Working' sheet
    for index, row in workings.iterrows():
        for col_num, value in enumerate(row, 1):
            working_sheet.cell(row=start_row_working, column=col_num, value=value)
        start_row_working += 1

    # Get or create the 'Reconciled' sheet in the existing workbook
    reconciled_sheet = existing_workbook.get_sheet_by_name('Reconciled') if 'Reconciled' in existing_workbook.sheetnames else existing_workbook.create_sheet(title='Reconciled')

    # Starting row number for the data in the 'Reconciled' sheet
    start_row_reconciled = 7

    # Write column headers to the first row of the 'Reconciled' sheet
    for col_num, header in enumerate(reconciled.columns, 1):
        reconciled_sheet.cell(row=start_row_reconciled - 1, column=col_num, value=header)

    # Write the 'Reconciled' DataFrame to the specified range in the 'Reconciled' sheet
    for index, row in reconciled.iterrows():
        for col_num, value in enumerate(row, 1):
            reconciled_sheet.cell(row=start_row_reconciled, column=col_num, value=value)
        start_row_reconciled += 1

    # Get or create the 'Unreconciled' sheet in the existing workbook
    unreconciled_sheet = existing_workbook.get_sheet_by_name('Non-reconciled') if 'Non-reconciled' in existing_workbook.sheetnames else existing_workbook.create_sheet(title='Non-reconciled')

    # Starting row number for the data in the 'Unreconciled' sheet
    start_row_unreconciled = 7

    # Write column headers to the first row of the 'Unreconciled' sheet
    for col_num, header in enumerate(unreconciled.columns, 1):
        unreconciled_sheet.cell(row=start_row_unreconciled - 1, column=col_num, value=header)

    # Write the 'Unreconciled' DataFrame to the specified range in the 'Unreconciled' sheet
    for index, row in unreconciled.iterrows():
        for col_num, value in enumerate(row, 1):
            unreconciled_sheet.cell(row=start_row_unreconciled, column=col_num, value=value)
        start_row_unreconciled += 1

    # Save the changes to the existing workbook
    existing_workbook['Sheet1'].title = 'NewTemplate'
    #print("def 5")

    existing_workbook.save(existing_workbook_path)



    # Calculate the sums
    length_cashbook= len(cashbook)
    length_cashreconciled= len(cash_reconciled)
    length_non_reconciled = length_cashbook-length_cashreconciled
    success_rate= (length_cashreconciled/length_cashbook)*100
    reconciled_sums = cash_reconciled['Un credited Items'].sum()
    reconciled_sums1 = cash_nonreconciled['Un credited Items'].sum()

    #cash_reconciled.to_csv("Rec.csv")

    results_df = pd.DataFrame({
        'Count Cashbook': [length_cashbook],
        'Count Cash Reconciled': [length_cashreconciled],
        'Count Non-Reconciled': [length_non_reconciled],
        'Percentage Success Rate': [f"{success_rate:.2f}%"],
        'Reconciled sums': [reconciled_sums],
        'Unreconciled sums': [reconciled_sums1]
    })






    # Open the existing workbook with openpyxl
    existing_workbook = openpyxl.load_workbook(existing_workbook_path)

    # Create a new sheet called 'Summary'
    summary_sheet = existing_workbook.create_sheet('Summary')

    # Write the 'results_df' DataFrame to the 'Summary' sheet
    summary_sheet.append(results_df.columns.tolist())

    for index, row in results_df.iterrows():
        summary_sheet.append(row.tolist())


    # Save the changes to the existing workbook
    existing_workbook.save(existing_workbook_path)




    # Replace these with your own email credentials and SMTP server details
    sender_email = "cynthiandululu@gmail.com"
    sender_password = "yruz vxwx jljp zqui"
    receiver_email = "jonagichohi@gmail.com"
    subject = f"RECONCILED/{file_name_without_extension} Summary"


    # Specify the directory where the file will be created
    directory_path = r"C:\Users\cynthia.mutisya\Downloads\Data (1)\Data\clean\RECONCILED"

    # Create the directory if it doesn't exist
    os.makedirs(directory_path, exist_ok=True)

    # Specify the filename with the correct extension (e.g., '.xlsx')
    file_name = f"{file_name_without_extension}_Recon.xlsx"
    file_path = os.path.join(directory_path, file_name)


    # Create the email message
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject

    # Body of the email
    body = "Please find attached the summary results."

    # Convert DataFrame to HTML table
    html_table1 = results_df.to_html(index=False)
    html_table = html_table1.replace('<table', '<table style="text-align: center;"')

    # Attach the HTML table to the email
    message.attach(MIMEText(body + html_table, "html"))

    # Attach the file
    attachment = open(file_path, "rb")
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename={file_name}')
    message.attach(part)
    attachment.close()
    # Establish a connection to the SMTP server
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()

        # Login to the email account
        server.login(sender_email, sender_password)

        # Send the email
        server.sendmail(sender_email, receiver_email, message.as_string())

    print("Email sent successfully.")



Process_Files(r"C:\Users\cynthia.mutisya\Downloads\Data (1)\Data\clean\STANBIC UGX Bank.xlsx",r"C:\Users\cynthia.mutisya\Downloads\Data (1)\Data\clean\STANBIC UGX Cashbook.xlsx",r"C:\Users\cynthia.mutisya\Downloads\Data (1)\Data\clean\STANBIC UGX Previous.xlsx")
Process_Files(r"C:\Users\cynthia.mutisya\Downloads\Data (1)\Data\clean\DTB UGX Bank.xlsx",r"C:\Users\cynthia.mutisya\Downloads\Data (1)\Data\clean\DTB UGX Cashbook.xlsx",r"C:\Users\cynthia.mutisya\Downloads\Data (1)\Data\clean\DTB UGX Previous.xlsx")
Process_Files(r"C:\Users\cynthia.mutisya\Downloads\Data (1)\Data\clean\DTB USD Bank.xlsx", r"C:\Users\cynthia.mutisya\Downloads\Data (1)\Data\clean\DTB USD Cashbook.xlsx", r"C:\Users\cynthia.mutisya\Downloads\Data (1)\Data\clean\DTB USD Previous.xlsx")
