import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from itertools import combinations
import warnings
import warnings
warnings.filterwarnings("ignore")
warnings.filterwarnings("default")
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase  
from email import encoders
import os




#print("def1")
def1=None
def2=None


def Process_Files(def1,def2,def3):
    df_bank=pd.read_excel(def1)

    pd.set_option('display.float_format',lambda x : '{:,.2f}'.format(x))
    # Use the existing file name as-is
    base_filename = os.path.basename(def1)

    # Extract the file name without extension
    file_name_without_extension, _ = os.path.splitext(base_filename)


    file_name_without_extension
    full_name=f"RECONCILED/{file_name_without_extension}_Recon.xlsx"
    print(full_name)

    df_bank.columns = df_bank.columns.str.strip()
    df_bank.dropna(subset="Transaction Date",inplace=True)
    df_bank=df_bank.rename(columns={'Transaction Date':'Date'})
    df_bank=df_bank.rename(columns={'Transaction Type':'VoucherN°'})
    df_bank=df_bank.rename(columns={'Transaction Details':'Description'})
    df_bank=df_bank.rename(columns={'Debits':'Direct Debits'})
    df_bank=df_bank.rename(columns={'Credits':'Un receipted Items'})
    df_bank.dropna(subset="Running Balance",inplace=True)
    df_bank=df_bank[['Date', 'VoucherN°', 'Description','Direct Debits', 'Un receipted Items']]
    df_bank['Index']=range(1,len(df_bank)+1)
    print(df_bank.shape)
    cashbook=pd.read_excel(def2)
    cashbook.dropna(subset="Journal Name",inplace=True)
    cashbook=cashbook.rename(columns={'Transaction Date':'Date'})
    cashbook=cashbook.rename(columns={'Reference Date':'Cheque N°'})
    cashbook=cashbook.rename(columns={'Line Description':'Description'})

    if "USD" in file_name_without_extension:
        Opening_balance= 220153.35
        float_columns = ['Debit FC', 'Credit FC']
        cashbook[float_columns] = cashbook[float_columns].astype(float)
        Total_Debit_FC = cashbook['Debit FC'].sum().round(2)
        Total_Credit_FC = cashbook['Credit FC'].sum().round(2)
        closing_balance_FC=(Opening_balance+Total_Debit_FC-Total_Credit_FC).round(2)
        print(closing_balance_FC)
        print(Total_Debit_FC)
        print(Total_Credit_FC)
        cashbook=cashbook.rename(columns={'Debit FC':'Un credited Items'})
        cashbook=cashbook.rename(columns={'Credit FC':'Un Paid Items'})
    else:
        cashbook=cashbook.rename(columns={'Debit LC':'Un credited Items'})
        cashbook=cashbook.rename(columns={'Credit LC':'Un Paid Items'})    
    cashbook=cashbook[['Date', 'Cheque N°', 'Description','Un credited Items', 'Un Paid Items']]
    cashbook['IndexCash']=range(1,len(cashbook)+1)
    print(cashbook.shape)
    
    #print("def2")
    previous_workings=pd.read_excel(def3)
    previous_workings.columns = previous_workings.columns.str.strip()
    previous_workings = previous_workings.dropna(subset=['Matching'])
    previous_workings=previous_workings[['Date', 'Cheque N°', 'Description','Direct Debits','Un receipted Items','Un credited Items', 'Un Paid Items','Matching']]
    
    print(previous_workings.dtypes)
    print(previous_workings.shape)
    #join the datasets
    #workings = pd.concat([cashbook, df_bank, previous_workings], axis=1, ignore_index=True)
    workings=pd.concat([cashbook,df_bank],ignore_index=True)
    workings=workings[['Date', 'VoucherN°', 'Cheque N°', 'Description','Direct Debits', 'Un receipted Items', 'Un credited Items','Un Paid Items','Index','IndexCash']]
    num_columns=['Direct Debits', 'Un receipted Items', 'Un credited Items','Un Paid Items']
    for col in num_columns:
        workings[col]=workings[col].astype(str)
        workings[col] = pd.to_numeric(workings[col].replace('-', pd.NaT).str.replace(',', ''), errors='coerce').fillna(0)

    workings['Matching']=workings['Direct Debits']+workings['Un Paid Items']+workings['Un credited Items']+workings['Un receipted Items']
    #workings.dropna(subset='Description',inplace=True)
    workings=workings.sort_values(by='Matching',ascending=False)
    workings.reset_index(drop=True,inplace=True)
    workings['Description'] = workings['Description'].str.strip()
    print(workings.shape)

    #workings=workings[workings['Matching']>0]
    workings['Total Index'] = workings[['Index', 'IndexCash']].sum(axis=1)
    print(workings.shape)

   
   

    workings=pd.concat([workings,previous_workings],ignore_index=True)


    similarity_threshold = 0

    def calculate_all_similarities():
        # Calculate similarities without parallel processing and batch processing
        similarities = []
        for i in range(len(workings['Description'])):
            for j in range(i + 1, len(workings['Description'])):
                similarity = fuzz.token_set_ratio(workings['Description'].iloc[i], workings['Description'].iloc[j])
                if similarity > similarity_threshold:
                    result = {
                        'Similarity': similarity,
                        'Index1': workings['Index'].iloc[i],
                        'IndexCash1': workings['IndexCash'].iloc[i],
                        'Row1': workings['Description'].iloc[i],
                        'Matching1': workings['Matching'].iloc[i],
                        'Date1': workings['Date'].iloc[i],
                        'VoucherN°1': workings['VoucherN°'].iloc[i],
                        'Cheque N°1': workings['Cheque N°'].iloc[i],
                        'Direct Debits1': workings['Direct Debits'].iloc[i],
                        'Un receipted Items1': workings['Un receipted Items'].iloc[i],
                        'Un credited Items1': workings['Un credited Items'].iloc[i],
                        'Un Paid Items1': workings['Un Paid Items'].iloc[i],
                        'Index2': workings['Index'].iloc[j],
                        'IndexCash2': workings['IndexCash'].iloc[j],
                        'Row2': workings['Description'].iloc[j],
                        'Matching2': workings['Matching'].iloc[j],
                        'Date2': workings['Date'].iloc[j],
                        'VoucherN°2': workings['VoucherN°'].iloc[j],
                        'Cheque N°2': workings['Cheque N°'].iloc[j],
                        'Direct Debits2': workings['Direct Debits'].iloc[j],
                        'Un receipted Items2': workings['Un receipted Items'].iloc[j],
                        'Un credited Items2': workings['Un credited Items'].iloc[j],
                        'Un Paid Items2': workings['Un Paid Items'].iloc[j],
                    }
                    similarities.append(result)

        return similarities





    # Call the function to calculate all similarities
    similarities = calculate_all_similarities()

    # Filter out None values (no similarity)
    similar_results = [result for result in similarities if result is not None]

    # Create a DataFrame from the list
    result_df = pd.DataFrame(similar_results)



    result_df.to_csv("lfg.csv")
    
    filtered_df = result_df.loc[(result_df['Matching1'] == result_df['Matching2'])]


    filtered_df.to_csv("files.csv")
    filtered_df=filtered_df.loc[(filtered_df['IndexCash1'] != filtered_df['IndexCash2'])&(filtered_df['Index1'] != filtered_df['Index2'])&(filtered_df['Direct Debits1']==filtered_df['Un Paid Items2'])&(filtered_df['Un receipted Items1']==filtered_df['Un credited Items2'])&(filtered_df['Un Paid Items1']==filtered_df['Direct Debits2'])&(filtered_df['Un credited Items1']==filtered_df['Un receipted Items2'])]
    filtered_df.to_csv('filtered.csv',index=False)


    #filtered_df=filtered_df.loc[(result_df['Direct Debits1']!=result_df['Direct Debits2'])&(result_df['Unreceipted Items1']!=result_df['Unreceipted Items2'])&(result_df['Un Paid Items1']!=result_df['Un Paid Items2'])&(result_df['Un credited Items1']!=result_df['Un credited Items2'])]

    # Separate columns based on 'Row1' and 'Row2'
    result_df1 = filtered_df[['Similarity', 'Row1', 'Matching1', 'Date1', 'VoucherN°1', 'Cheque N°1', 'Direct Debits1', 'Un receipted Items1', 'Un credited Items1', 'Un Paid Items1','Index1','IndexCash1']]
    result_df2 = filtered_df[['Similarity', 'Row2', 'Matching2', 'Date2', 'VoucherN°2', 'Cheque N°2', 'Direct Debits2', 'Un receipted Items2', 'Un credited Items2', 'Un Paid Items2','Index2','IndexCash2']]
    alldf= pd.concat([result_df1,result_df2])
    filtered_df.to_csv("alldf.csv")
    # Remove the '1' suffix from the column names in result_df1
    # Remove the '1' suffix from the column names in result_df1
    result_df = result_df1.rename(columns=lambda x: x.rstrip('1'))

    # Remove the '2' suffix from the column names in result_df2
    result_df = result_df2.rename(columns=lambda x: x.rstrip('2'))
    matching_fuzzy = pd.concat([result_df, result_df])
    matching_fuzzy=matching_fuzzy.rename(columns={'Row':'Description'})
    matching_fuzzy['Total Index'] = matching_fuzzy[['Index', 'IndexCash']].sum(axis=1)

    #matching_fuzzy['Total Index']=matching_fuzzy['Index']+matching_fuzzy['IndexCash']
    # matching_fuzzy=matching_fuzzy[['Date', 'VoucherN°', 'Cheque N°', 'Description', 'Direct Debits',
    #        'Unreceipted Items', 'Un credited Items', 'Un Paid Items', 'Matching']]

    selected_columns = ['Date', 'VoucherN°', 'Cheque N°', 'Description', 'Direct Debits', 'Un receipted Items', 'Un credited Items', 'Un Paid Items','Index','Total Index','IndexCash','Matching']
    reconciled = matching_fuzzy[selected_columns]
    reconciled = reconciled[reconciled['Matching']>0]
    '''reconciled=pd.concat([matching_fuzzy,duplicate_rows1])'''
    unreconciled = workings[~workings['Description'].isin(reconciled['Description'])]
    unreconciled = unreconciled[selected_columns]
    
    reconciled = reconciled.drop_duplicates(subset=['IndexCash', 'Index'], keep='first')
    #selected_columns1 = ['Date', 'VoucherN°', 'Cheque N°', 'Description', 'Direct Debits', 'Unreceipted Items', 'Un credited Items', 'Un Paid Items','Matching']
    #reconciled = reconciled[selected_columns1]
    #unreconciled = unreconciled[selected_columns1]
    


    reconciled.to_csv("Reconciled.csv")
    unreconciled.to_csv("UnReconciled.csv")

    
    df_bank1=pd.read_excel(def1)
    df_bank1.columns = df_bank1.columns.str.strip()


    # Convert the 'Transaction Transaction Date' column to Transaction Datetime
    df_bank1['Transaction Date'] = pd.to_datetime(df_bank1['Transaction Date'], errors='coerce')

    # Sort the DataFrame by the ' Date' column
    df_bank1 = df_bank1.sort_values(by='Transaction Date',ascending=True)

    # If you want to reset the index after sorting
    df_bank1 = df_bank1.reset_index(drop=True)
    df_bank1.dropna(subset='Running Balance',inplace=True)

    # Print the sorted DataFrame

    running_balance_last_row = df_bank1['Running Balance'].iloc[-1]
    if isinstance(running_balance_last_row, str) and ',' in running_balance_last_row:
        running_balance_last_row = float(running_balance_last_row.replace(',', ''))
    elif isinstance(running_balance_last_row, (int, float)):
        running_balance_last_row = float(running_balance_last_row)
    cashbook1=pd.read_excel(def2)
    cashbook1.dropna(subset='Running Total',inplace=True)
    running_cashbook_last_row = cashbook1['Running Total'].iloc[-1]
    running_cashbook_last_row=running_cashbook_last_row
    # Assuming unreconciled is your non-reconciled dataframe
    '''unreconciled=unreconciled.dropna(subset='Description')

    sum_row['Date'] = 'Sub Total'

    #unreconciled.loc[len(unreconciled)] = sum_row
    unreconciled.reset_index(drop=True, inplace=True)'''
    # Convert sum_row to a DataFrame with a single row
    sum_row = unreconciled.select_dtypes(include='float').sum()

    sum_row_df = pd.DataFrame([sum_row], columns=sum_row.index)

    # Concatenate unreconciled and sum_row_df
    unreconciled = pd.concat([unreconciled, sum_row_df], ignore_index=True)

    # Set 'Date' to 'Sub Total' in the last row
    unreconciled.loc[len(unreconciled), 'Date'] = 'Sub Total'
    unreconciled=unreconciled.dropna(subset='Matching')

    unreconciled.reset_index(drop=True, inplace=True)
    cash_reconciled = cashbook[cashbook['Description'].isin(reconciled['Description'])]
    cash_nonreconciled = cashbook[~cashbook['Description'].isin(reconciled['Description'])]



    sum1=unreconciled['Un credited Items'].iloc[-1]
    sum2=unreconciled['Direct Debits'].iloc[-1]
    sum3=unreconciled['Unreceipted Items'].iloc[-1]
    sum4=unreconciled['Un Paid Items'].iloc[-1]

    # Load the workbook
    workbook = openpyxl.load_workbook("8E3E8A00.xlsx")

    # Select the desired worksheet (replace "Sheet1" with the actual sheet name)
    worksheet = workbook["Sheet1"]

    # Write sum1 to cell C8
    
    if "USD" in full_name:
        worksheet["C10"] = closing_balance_FC

    else:
        worksheet["C10"] = running_cashbook_last_row

    #worksheet["C10"] = running_cashbook_last_row
    worksheet["D8"] = running_balance_last_row
    worksheet["D13"] = sum1
    worksheet["C14"] = sum2
    worksheet["C17"] = sum3
    worksheet["D18"] = sum4
    #worksheet["C24"]=worksheet["C10"]-worksheet["C14"]+worksheet["C17"]
    #worksheet["C24"]=worksheet["C10"]-worksheet["C14"]+worksheet["C17"]

    # Save the changes
    workbook.save(full_name)


    # Load an existing workbook
    existing_workbook_path = full_name
    existing_workbook = load_workbook(existing_workbook_path)


    # Get or create the 'Working' sheet in the existing workbook
    working_sheet = existing_workbook.get_sheet_by_name('Working') if 'Working' in existing_workbook.sheetnames else existing_workbook.create_sheet(title='Working')
    working_sheet.sheet_properties.tabColor = 'FFA500'  # Hex color code for orange

    # Starting row number for the data in the 'Working' sheet
    start_row_working = 7

    # Write column headers to the first row of the 'Working' sheet
    for col_num, header in enumerate(workings.columns, 1):
        working_sheet.cell(row=start_row_working - 1, column=col_num, value=header)

    # Write the 'Working' DataFrame to the specified range in the 'Working' sheet
    for index, row in workings.iterrows():
        for col_num, value in enumerate(row, 1):
            working_sheet.cell(row=start_row_working, column=col_num, value=value)
        start_row_working += 1

    # Get or create the 'Reconciled' sheet in the existing workbook
    reconciled_sheet = existing_workbook.get_sheet_by_name('Reconciled') if 'Reconciled' in existing_workbook.sheetnames else existing_workbook.create_sheet(title='Reconciled')

    # Starting row number for the data in the 'Reconciled' sheet
    start_row_reconciled = 7

    # Write column headers to the first row of the 'Reconciled' sheet
    for col_num, header in enumerate(reconciled.columns, 1):
        reconciled_sheet.cell(row=start_row_reconciled - 1, column=col_num, value=header)

    # Write the 'Reconciled' DataFrame to the specified range in the 'Reconciled' sheet
    for index, row in reconciled.iterrows():
        for col_num, value in enumerate(row, 1):
            reconciled_sheet.cell(row=start_row_reconciled, column=col_num, value=value)
        start_row_reconciled += 1

    # Get or create the 'Unreconciled' sheet in the existing workbook
    unreconciled_sheet = existing_workbook.get_sheet_by_name('Non-reconciled') if 'Non-reconciled' in existing_workbook.sheetnames else existing_workbook.create_sheet(title='Non-reconciled')

    # Starting row number for the data in the 'Unreconciled' sheet
    start_row_unreconciled = 7

    # Write column headers to the first row of the 'Unreconciled' sheet
    for col_num, header in enumerate(unreconciled.columns, 1):
        unreconciled_sheet.cell(row=start_row_unreconciled - 1, column=col_num, value=header)

    # Write the 'Unreconciled' DataFrame to the specified range in the 'Unreconciled' sheet
    for index, row in unreconciled.iterrows():
        for col_num, value in enumerate(row, 1):
            unreconciled_sheet.cell(row=start_row_unreconciled, column=col_num, value=value)
        start_row_unreconciled += 1

    # Save the changes to the existing workbook
    existing_workbook['Sheet1'].title = 'NewTemplate'
    #print("def 5")

    existing_workbook.save(existing_workbook_path)



    # Calculate the sums
    length_cashbook= len(cashbook)
    length_cashreconciled= len(cash_reconciled)
    length_non_reconciled = length_cashbook-length_cashreconciled
    success_rate= (length_cashreconciled/length_cashbook)*100
    reconciled_sums = cash_reconciled['Un credited Items'].sum()
    reconciled_sums1 = cash_nonreconciled['Un credited Items'].sum()

    #cash_reconciled.to_csv("Rec.csv")

    results_df = pd.DataFrame({
        'Count Cashbook': [length_cashbook],
        'Count Cash Reconciled': [length_cashreconciled],
        'Count Non-Reconciled': [length_non_reconciled],
        'Percentage Success Rate': [f"{success_rate:.2f}%"],
        'Reconciled sums': [reconciled_sums],
        'Unreconciled sums': [reconciled_sums1]
    })






    # Open the existing workbook with openpyxl
    existing_workbook = openpyxl.load_workbook(existing_workbook_path)

    # Create a new sheet called 'Summary'
    summary_sheet = existing_workbook.create_sheet('Summary')

    # Write the 'results_df' DataFrame to the 'Summary' sheet
    summary_sheet.append(results_df.columns.tolist())

    for index, row in results_df.iterrows():
        summary_sheet.append(row.tolist())


    # Save the changes to the existing workbook
    existing_workbook.save(existing_workbook_path)




    # Replace these with your own email credentials and SMTP server details
    sender_email = "cynthiandululu@gmail.com"
    sender_password = "yruz vxwx jljp zqui"
    receiver_email = "jonagichohi@gmail.com"
    subject = f"RECONCILED/{file_name_without_extension} Summary"


    # Specify the directory where the file will be created
    directory_path = 'C:\\Users\\robot.nation25\\Documents\\Uipath\\Reconciliation\\Data\\clean\\RECONCILED'

    # Create the directory if it doesn't exist
    os.makedirs(directory_path, exist_ok=True)

    # Specify the filename with the correct extension (e.g., '.xlsx')
    file_name = f"{file_name_without_extension}_Recon.xlsx"
    file_path = os.path.join(directory_path, file_name)


    # Create the email message
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject

    # Body of the email
    body = "Please find attached the summary results."

    # Convert DataFrame to HTML table
    html_table1 = results_df.to_html(index=False)
    html_table = html_table1.replace('<table', '<table style="text-align: center;"')

    # Attach the HTML table to the email
    message.attach(MIMEText(body + html_table, "html"))

    # Attach the file
    attachment = open(file_path, "rb")
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename={file_name}')
    message.attach(part)
    attachment.close()
    # Establish a connection to the SMTP server
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()

        # Login to the email account
        server.login(sender_email, sender_password)

        # Send the email
        server.sendmail(sender_email, receiver_email, message.as_string())

    print("Email sent successfully.")



Process_Files(r"C:\Users\robot.nation25\Documents\Uipath\Reconciliation\Data\clean\STANBIC UGX Bank.xlsx",r"C:\Users\robot.nation25\Documents\Uipath\Reconciliation\Data\clean\STANBIC UGX Cashbook.xlsx",r"C:\Users\robot.nation25\Documents\Uipath\Reconciliation\Data\clean\STANBIC UGX Previous.xlsx")
#Process_Files(r"C:\Users\robot.nation25\Documents\Uipath\Reconciliation\Data\clean\DTB UGX Bank.xlsx",r"C:\Users\robot.nation25\Documents\Uipath\Reconciliation\Data\clean\DTB UGX Cashbook.xlsx",r"C:\Users\robot.nation25\Documents\Uipath\Reconciliation\Data\clean\DTB UGX Previous.xlsx")
#Process_Files(r"C:\Users\robot.nation25\Documents\Uipath\Reconciliation\Data\clean\DTB USD Bank.xlsx",r"C:\Users\robot.nation25\Documents\Uipath\Reconciliation\Data\clean\DTB USD Cashbook.xlsx",r"C:\Users\robot.nation25\Documents\Uipath\Reconciliation\Data\clean\DTB USD Previous.xlsx")
